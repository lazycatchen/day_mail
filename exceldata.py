# _*_ coding: utf-8 _*_
import xlrd
import os
import time
import xlwt
from openpyxl import load_workbook,Workbook
from xlrd import open_workbook
from xlutils.copy import copy

def forder(filename): #返回字典键
    chx=''
    temp=['内蒙','靖边','干北','诺木洪','新疆','河北','江苏','敦煌','共和','山东','尧生','光热']  #re1
    for ch in temp:
         if  ch in filename:
             chx=ch
    return chx

def custom_key(word):
   numbers = []
   orderstr=[ '内蒙', '靖边','干北','诺木洪','新疆','河北','江苏','敦煌','共和','山东','尧生','光热']  #re1 #以order顺序读取文件
   temp=[i for i, word1 in enumerate(orderstr) if word1 in word]
   numbers.append(temp)
   return numbers

def exceltable(path,datetable):
    g = os.walk(path)
    rb = open_workbook('F:\\0\\py_ribao\\py_save\\py日报模板.xls', formatting_info=True)
   # wb1 = copy(rb)              #读取模板内容，模板相当于本地数据库存储的是月、年总的信息，需要每月调整一次
   # ws = wb1.get_sheet(0)
    pathtable='F:\\0\\py_ribao\\py_save\\table.xls'
    for path,dir_list,file_list in g:
        file_list.sort(key=custom_key)   #指定场站先后顺序!
        listdata=[]
        xls_result= xlwt.Workbook()
        sht1 = xls_result.add_sheet('Sheet1',cell_overwrite_ok=True)   #写入sheet1文件
        j=0
        if int(datetable[-2:])<10:
            nowdate=int(datetable[-1])-1     #读取不同文件中的日期。小于10日即1-9，为1位数，其余为2位数
        else:
            nowdate=int(datetable[-2:])-1
        for file_name in file_list:   #遍历文件夹，依次读取文件
            data = xlrd.open_workbook(os.path.join(path, file_name))
            names = data.sheet_names()
            #table = data.sheets()[len(names)-1] #读取文件至table，大部分场站把当天文件放到了最后一张sheet，新疆、青海、陕西自行命名

            ch=forder(file_name)
            if ch=='新疆':               #新疆隐藏了一张sheet
                table=data.sheets()[nowdate-1+60]
            elif ch=='诺木洪' or ch=='共和'or ch=='光热':   #r1#青海按日期命名
                date1=str(nowdate)
                table=data.sheet_by_name(date1)
            elif ch=='靖边':                #靖边随缘命名
                table=data.sheets()[nowdate-1]
            else:
                table=data.sheets()[nowdate-1]

            templist=[]
            for (temp,temp2) in zip(table._cell_types ,table._cell_values) :  #读取文件中数据
                if temp.count(2)==10 or temp.count(2)==11 or temp.count(2)==13:   #根据其独特的数据结构，只要是数字其_cell_types为2
                    j=j+1 #excel对应的行
                    listdata.append(temp2)
                    for i,data in enumerate(temp2):  #遍历存入
                        sht1.write(j,0,ch)
                        #ws.write(j,0,ch)   #
                        sht1.write(j,i+1,data)
                       # ws.write(j,i+1,data)  #
                        templist.append(temp2) #单表数据提取
        sht1.write(0,0,int(datetable[-4:-2]))   #存储到模板
        sht1.write(0,1,int(datetable[-2:])-1)##日期
        xls_result.save(pathtable)
        wb = Workbook()
        wb.save(datetable+'.xls')
        time.sleep(1)
        xls_result.save(datetable+'.xls')
        pathtable1='F:\\0\\py_ribao\\py_save\\'+str(nowdate)
        return pathtable1



